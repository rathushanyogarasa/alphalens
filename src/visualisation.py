"""Visualisation and reporting module for AlphaLens.

Generates all diagnostic and performance plots and exports outputs to
Excel-ready tables.  All functions are self-contained — pass in a
DataFrame of returns or a model object and receive a saved figure.

Generated plots
~~~~~~~~~~~~~~~
* Equity curve
* Drawdown
* Signal vs return scatter
* Feature importance (bar chart)
* Confusion matrix
* Return distribution (histogram)
* Rolling Sharpe
* Signal decay (IC at multiple horizons)

Export formats
~~~~~~~~~~~~~~
* PNG figures (saved to ``config.PLOTS_DIR``)
* Excel workbook (saved to ``config.RESULTS_DIR``)

Usage::

    from src.visualisation import AlphaLensVisualiser
    vis = AlphaLensVisualiser()
    vis.generate_full_report(ticker="NVDA", result=recommendation_result)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
import config

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Shared plot style
# ---------------------------------------------------------------------------

_COLOURS = {
    "primary":   "#1a73e8",
    "positive":  "#34a853",
    "negative":  "#ea4335",
    "neutral":   "#fbbc04",
    "background":"#f8f9fa",
    "grid":      "#dadce0",
    "text":      "#202124",
}


def _apply_style(ax) -> None:
    """Apply a consistent style to a matplotlib Axes object."""
    ax.set_facecolor(_COLOURS["background"])
    ax.figure.set_facecolor("white")
    ax.grid(True, color=_COLOURS["grid"], linewidth=0.7, alpha=0.7)
    ax.tick_params(colors=_COLOURS["text"], labelsize=9)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color(_COLOURS["grid"])
    ax.spines["bottom"].set_color(_COLOURS["grid"])


def _save_figure(fig, name: str) -> Path:
    """Save figure to ``config.PLOTS_DIR`` and close it."""
    path = config.PLOTS_DIR / f"{name}.png"
    fig.savefig(path, dpi=150, bbox_inches="tight")
    logger.info("Plot saved: %s", path)
    try:
        import matplotlib.pyplot as plt
        plt.close(fig)
    except Exception:
        pass
    return path


# ---------------------------------------------------------------------------
# AlphaLensVisualiser
# ---------------------------------------------------------------------------


class AlphaLensVisualiser:
    """Generates all AlphaLens diagnostic and reporting plots.

    Args:
        plots_dir: Override for plot output directory.
    """

    def __init__(self, plots_dir: Path | None = None) -> None:
        self._plots_dir = plots_dir or config.PLOTS_DIR
        self._plots_dir.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # 1. Equity curve
    # ------------------------------------------------------------------

    def plot_equity_curve(
        self,
        returns: pd.Series,
        benchmark: pd.Series | None = None,
        label: str = "Strategy",
        tag: str = "",
    ) -> Path:
        """Plot cumulative return equity curve.

        Args:
            returns:   Daily strategy return series.
            benchmark: Optional daily benchmark return series.
            label:     Strategy legend label.
            tag:       Optional filename suffix (e.g. ticker).

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt

        cum = (1 + returns.fillna(0)).cumprod() - 1

        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(cum.index, cum.values * 100, color=_COLOURS["primary"],
                linewidth=1.5, label=label)

        if benchmark is not None:
            cum_b = (1 + benchmark.fillna(0)).cumprod() - 1
            ax.plot(cum_b.index, cum_b.values * 100, color=_COLOURS["neutral"],
                    linewidth=1.2, linestyle="--", label="Benchmark")

        ax.axhline(0, color=_COLOURS["grid"], linewidth=1)
        ax.fill_between(cum.index, cum.values * 100, 0,
                        where=(cum.values >= 0),
                        alpha=0.12, color=_COLOURS["positive"])
        ax.fill_between(cum.index, cum.values * 100, 0,
                        where=(cum.values < 0),
                        alpha=0.12, color=_COLOURS["negative"])

        ax.set_title("Equity Curve — Cumulative Return", fontsize=12, fontweight="bold")
        ax.set_ylabel("Cumulative Return (%)")
        ax.legend(fontsize=9)
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"equity_curve{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 2. Drawdown
    # ------------------------------------------------------------------

    def plot_drawdown(
        self, returns: pd.Series, label: str = "Strategy", tag: str = ""
    ) -> Path:
        """Plot drawdown from peak over time.

        Args:
            returns: Daily strategy return series.
            label:   Plot label.
            tag:     Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt

        cum = (1 + returns.fillna(0)).cumprod()
        roll_max = cum.cummax()
        drawdown = (cum - roll_max) / roll_max * 100

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.fill_between(drawdown.index, drawdown.values, 0,
                        color=_COLOURS["negative"], alpha=0.5, label=label)
        ax.plot(drawdown.index, drawdown.values, color=_COLOURS["negative"], linewidth=0.8)
        ax.axhline(0, color=_COLOURS["grid"], linewidth=1)

        max_dd = float(drawdown.min())
        ax.annotate(
            f"Max DD: {max_dd:.1f}%",
            xy=(drawdown.idxmin(), max_dd),
            xytext=(drawdown.idxmin(), max_dd - 3),
            fontsize=8, color=_COLOURS["negative"],
            arrowprops={"arrowstyle": "->", "color": _COLOURS["negative"]},
        )

        ax.set_title("Drawdown (%)", fontsize=12, fontweight="bold")
        ax.set_ylabel("Drawdown (%)")
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"drawdown{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 3. Signal vs return scatter
    # ------------------------------------------------------------------

    def plot_signal_scatter(
        self,
        signals: pd.Series,
        returns: pd.Series,
        tag: str = "",
    ) -> Path:
        """Scatter plot of signal scores vs forward returns.

        Args:
            signals: Composite signal score series.
            returns: Corresponding forward returns.
            tag:     Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt
        from scipy import stats

        df = pd.DataFrame({"signal": signals, "return": returns}).dropna()
        if df.empty:
            logger.warning("Empty signal/return data for scatter plot")
            return config.PLOTS_DIR / "signal_scatter_empty.png"

        colours_mapped = np.where(df["return"] > 0, _COLOURS["positive"], _COLOURS["negative"])

        fig, ax = plt.subplots(figsize=(8, 6))
        ax.scatter(df["signal"], df["return"] * 100, c=colours_mapped,
                   alpha=0.5, s=18, edgecolors="none")

        # Regression line
        if len(df) > 5:
            slope, intercept, r_val, p_val, _ = stats.linregress(df["signal"], df["return"] * 100)
            x_line = np.linspace(df["signal"].min(), df["signal"].max(), 100)
            ax.plot(x_line, slope * x_line + intercept,
                    color=_COLOURS["primary"], linewidth=1.5, linestyle="--",
                    label=f"R²={r_val**2:.3f}  p={p_val:.3f}")
            ax.legend(fontsize=9)

        ax.axhline(0, color=_COLOURS["grid"], linewidth=1)
        ax.axvline(0, color=_COLOURS["grid"], linewidth=1)
        ax.set_xlabel("Signal Score")
        ax.set_ylabel("Forward Return (%)")
        ax.set_title("Signal Score vs Forward Return", fontsize=12, fontweight="bold")
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"signal_scatter{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 4. Feature importance
    # ------------------------------------------------------------------

    def plot_feature_importance(
        self, importances: pd.Series, top_n: int = 25, tag: str = ""
    ) -> Path:
        """Horizontal bar chart of top feature importances.

        Args:
            importances: Series with feature names as index, sorted descending.
            top_n:       Number of features to display.
            tag:         Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt

        top = importances.head(top_n)
        if top.empty:
            logger.warning("Empty importances for feature importance plot")
            return config.PLOTS_DIR / "feature_importance_empty.png"

        fig_height = max(4, top_n * 0.3)
        fig, ax = plt.subplots(figsize=(9, fig_height))

        colours = [
            _COLOURS["positive"] if "sent_" in k or "macro_" in k
            else _COLOURS["primary"] if "market_" in k
            else _COLOURS["neutral"]
            for k in top.index
        ]
        bars = ax.barh(top.index[::-1], top.values[::-1] * 100, color=colours[::-1])

        ax.set_xlabel("Importance (%)")
        ax.set_title(f"Top {top_n} Feature Importances", fontsize=12, fontweight="bold")
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"feature_importance{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 5. Confusion matrix
    # ------------------------------------------------------------------

    def plot_confusion_matrix(
        self,
        y_true: list | np.ndarray,
        y_pred: list | np.ndarray,
        labels: list[str] | None = None,
        tag: str = "",
    ) -> Path:
        """Normalised confusion matrix heatmap.

        Args:
            y_true:  True class labels.
            y_pred:  Predicted class labels.
            labels:  Class label strings (default: negative/neutral/positive).
            tag:     Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt
        from sklearn.metrics import confusion_matrix
        import seaborn as sns

        labels = labels or ["negative", "neutral", "positive"]
        cm = confusion_matrix(y_true, y_pred, labels=labels)
        cm_norm = cm.astype(float) / cm.sum(axis=1, keepdims=True)

        fig, ax = plt.subplots(figsize=(7, 5))
        sns.heatmap(
            cm_norm, annot=True, fmt=".2f",
            xticklabels=labels, yticklabels=labels,
            cmap="Blues", linewidths=0.5, ax=ax, vmin=0, vmax=1,
        )
        ax.set_xlabel("Predicted")
        ax.set_ylabel("Actual")
        ax.set_title("Normalised Confusion Matrix", fontsize=12, fontweight="bold")
        fig.tight_layout()
        return _save_figure(fig, f"confusion_matrix{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 6. Return distribution
    # ------------------------------------------------------------------

    def plot_return_distribution(
        self, returns: pd.Series, tag: str = ""
    ) -> Path:
        """Histogram of daily strategy returns.

        Args:
            returns: Daily return series.
            tag:     Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt
        from scipy.stats import norm

        data = returns.dropna() * 100
        if data.empty:
            return config.PLOTS_DIR / "return_dist_empty.png"

        fig, ax = plt.subplots(figsize=(9, 5))
        ax.hist(data, bins=50, color=_COLOURS["primary"], alpha=0.7,
                edgecolor="white", linewidth=0.4, density=True, label="Returns")

        # Normal overlay
        mu, sigma = data.mean(), data.std()
        x = np.linspace(data.min(), data.max(), 200)
        ax.plot(x, norm.pdf(x, mu, sigma), color=_COLOURS["negative"],
                linewidth=1.5, label=f"Normal(μ={mu:.2f}%, σ={sigma:.2f}%)")

        ax.axvline(0, color=_COLOURS["grid"], linewidth=1, linestyle="--")
        ax.axvline(mu, color=_COLOURS["neutral"], linewidth=1.5,
                   linestyle="--", label=f"Mean={mu:.2f}%")

        ax.set_xlabel("Daily Return (%)")
        ax.set_ylabel("Density")
        ax.set_title("Return Distribution", fontsize=12, fontweight="bold")
        ax.legend(fontsize=8)
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"return_distribution{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 7. Rolling Sharpe
    # ------------------------------------------------------------------

    def plot_rolling_sharpe(
        self,
        returns: pd.Series,
        window: int = 60,
        risk_free: float | None = None,
        tag: str = "",
    ) -> Path:
        """Rolling Sharpe ratio over time.

        Args:
            returns:    Daily return series.
            window:     Rolling window in trading days (default 60).
            risk_free:  Annualised risk-free rate (default from config).
            tag:        Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt

        rf = risk_free if risk_free is not None else config.RISK_FREE_RATE
        rf_daily = rf / 252

        excess = returns.fillna(0) - rf_daily
        roll_mean = excess.rolling(window).mean() * 252
        roll_std = excess.rolling(window).std() * np.sqrt(252)
        rolling_sharpe = roll_mean / roll_std.replace(0, np.nan)

        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(rolling_sharpe.index, rolling_sharpe.values,
                color=_COLOURS["primary"], linewidth=1.3, label=f"Rolling {window}d Sharpe")
        ax.axhline(0, color=_COLOURS["grid"], linewidth=1)
        ax.axhline(1, color=_COLOURS["positive"], linewidth=1,
                   linestyle="--", alpha=0.7, label="Sharpe = 1")
        ax.axhline(-1, color=_COLOURS["negative"], linewidth=1,
                   linestyle="--", alpha=0.7, label="Sharpe = -1")

        ax.fill_between(rolling_sharpe.index, rolling_sharpe.values, 0,
                        where=(rolling_sharpe.values >= 0),
                        alpha=0.10, color=_COLOURS["positive"])
        ax.fill_between(rolling_sharpe.index, rolling_sharpe.values, 0,
                        where=(rolling_sharpe.values < 0),
                        alpha=0.10, color=_COLOURS["negative"])

        ax.set_ylabel("Sharpe Ratio")
        ax.set_title(f"Rolling {window}-Day Sharpe Ratio", fontsize=12, fontweight="bold")
        ax.legend(fontsize=8)
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"rolling_sharpe{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 8. Signal decay (IC at multiple horizons)
    # ------------------------------------------------------------------

    def plot_signal_decay(
        self,
        ic_decay: dict[int, float],
        tag: str = "",
    ) -> Path:
        """Bar chart of IC at multiple forward-return horizons.

        Args:
            ic_decay: Dict mapping lag (days) → IC value.
            tag:      Filename suffix.

        Returns:
            Path: Saved figure path.
        """
        import matplotlib.pyplot as plt

        if not ic_decay:
            return config.PLOTS_DIR / "signal_decay_empty.png"

        lags = sorted(ic_decay.keys())
        ics = [ic_decay[l] for l in lags]

        colours = [_COLOURS["positive"] if v > 0 else _COLOURS["negative"] for v in ics]

        fig, ax = plt.subplots(figsize=(8, 5))
        bars = ax.bar([str(l) for l in lags], ics, color=colours, alpha=0.8, width=0.5)
        ax.axhline(0, color=_COLOURS["text"], linewidth=0.8)

        for bar, val in zip(bars, ics):
            ax.text(bar.get_x() + bar.get_width() / 2, val + 0.001 * np.sign(val),
                    f"{val:.3f}", ha="center", va="bottom" if val > 0 else "top",
                    fontsize=8)

        ax.set_xlabel("Forward Return Lag (days)")
        ax.set_ylabel("Information Coefficient")
        ax.set_title("Signal Decay — IC at Multiple Horizons", fontsize=12, fontweight="bold")
        _apply_style(ax)
        fig.tight_layout()
        return _save_figure(fig, f"signal_decay{'_' + tag if tag else ''}")

    # ------------------------------------------------------------------
    # 9. Excel export
    # ------------------------------------------------------------------

    def export_to_excel(
        self,
        results: list,
        output_path: Path | None = None,
    ) -> Path:
        """Export recommendation results to an Excel workbook.

        Creates two sheets: **Summary** (portfolio overview) and
        **Details** (per-ticker metrics).

        Args:
            results:     List of :class:`~src.stock_engine.RecommendationResult`.
            output_path: Override output path.

        Returns:
            Path: Saved workbook path.
        """
        out_path = output_path or (config.RESULTS_DIR / "alphalens_report.xlsx")

        rows = []
        for r in results:
            row = {
                "Ticker":          r.ticker,
                "Recommendation":  r.recommendation,
                "Confidence":      r.confidence,
                "Signal Score":    r.signal_score,
                "Headlines":       r.headline_count,
                "Confident HLs":   r.confident_headline_count,
                "Lexicon Score":   r.lexicon_score,
                "Top Positive KW": ", ".join(r.top_positive_keywords[:3]),
                "Top Negative KW": ", ".join(r.top_negative_keywords[:3]),
                "Sources":         str(r.source_breakdown),
                "Macro Regime":    getattr(r, "macro_regime", ""),
                "Model Quality":   getattr(r, "model_quality_score", 0.0),
                "Drivers":         " | ".join(getattr(r, "drivers", [])),
                "Risk Flags":      " | ".join(getattr(r, "risk_flags", [])),
                "Generated At":    r.generated_at.strftime("%Y-%m-%d %H:%M UTC"),
                "Reasoning":       r.reasoning,
            }
            rows.append(row)

        if not rows:
            logger.warning("No results to export")
            return out_path

        df_summary = pd.DataFrame(rows)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"

            # Header style
            header_fill = PatternFill("solid", fgColor="1a73e8")
            header_font = Font(bold=True, color="FFFFFF", size=10)

            # Write header
            for col_idx, col_name in enumerate(df_summary.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Write data rows with conditional colour
            rec_colours = {"BUY": "c8e6c9", "SELL": "ffcdd2", "HOLD": "fff9c4"}
            for row_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    rec_val = df_summary.iloc[row_idx - 2]["Recommendation"]
                    bg = rec_colours.get(rec_val, "FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=bg)

            # Auto-width
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

            # Details sheet — reasoning
            ws2 = wb.create_sheet("Reasoning")
            ws2.column_dimensions["A"].width = 12
            ws2.column_dimensions["B"].width = 15
            ws2.column_dimensions["C"].width = 80
            ws2.append(["Ticker", "Recommendation", "Full Reasoning"])
            for _, row in df_summary.iterrows():
                ws2.append([row["Ticker"], row["Recommendation"], row["Reasoning"]])

            wb.save(out_path)
            logger.info("Excel report saved: %s", out_path)
        except ImportError:
            logger.warning("openpyxl not installed — saving CSV instead")
            out_path = out_path.with_suffix(".csv")
            df_summary.to_csv(out_path, index=False)

        return out_path

    # ------------------------------------------------------------------
    # 10. Full report
    # ------------------------------------------------------------------

    def generate_full_report(
        self,
        ticker: str,
        result,
        returns: pd.Series | None = None,
        ic_decay: dict[int, float] | None = None,
        feature_importances: pd.Series | None = None,
        y_true: list | None = None,
        y_pred: list | None = None,
    ) -> list[Path]:
        """Generate all available plots for a single ticker analysis.

        Args:
            ticker:              Equity ticker.
            result:              RecommendationResult.
            returns:             Daily strategy return series (optional).
            ic_decay:            IC decay dict (optional).
            feature_importances: Feature importance Series (optional).
            y_true:              True labels for confusion matrix (optional).
            y_pred:              Predicted labels (optional).

        Returns:
            list[Path]: List of all saved plot paths.
        """
        saved: list[Path] = []
        tag = ticker.lower()

        if returns is not None and not returns.empty:
            saved.append(self.plot_equity_curve(returns, tag=tag))
            saved.append(self.plot_drawdown(returns, tag=tag))
            saved.append(self.plot_return_distribution(returns, tag=tag))
            saved.append(self.plot_rolling_sharpe(returns, tag=tag))

        if ic_decay:
            saved.append(self.plot_signal_decay(ic_decay, tag=tag))

        if feature_importances is not None and not feature_importances.empty:
            saved.append(self.plot_feature_importance(feature_importances, tag=tag))

        if y_true is not None and y_pred is not None:
            saved.append(self.plot_confusion_matrix(y_true, y_pred, tag=tag))

        logger.info("[%s] Full report generated: %d plots", ticker, len(saved))
        return saved
